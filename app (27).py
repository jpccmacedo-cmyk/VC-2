import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell
from copy import copy
from datetime import datetime, date, timedelta
from io import BytesIO
import zipfile
import tempfile
import shutil
import os
import re
import gc
import time
import unicodedata
import traceback
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
st.set_page_config(page_title="Consolidador Gerencial CN", page_icon="📊", layout="wide")

NOME_ABA_GERENCIAL = "Gerencial"
COLUNA_INICIO_LIMPEZA = 13  # Coluna M
ORDEM_ABAS = ["COB", "CUI", "EDE", "NOB", "PVE", "SOB", "XAM"]
MESES_PT = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "março": 3, "abril": 4,
    "maio": 5, "junho": 6, "julho": 7, "agosto": 8, "setembro": 9,
    "outubro": 10, "novembro": 11, "dezembro": 12,
}

# ============================================================
# UTILITÁRIOS
# ============================================================

def remover_acentos(texto):
    texto = str(texto)
    texto = unicodedata.normalize("NFKD", texto)
    texto = texto.encode("ASCII", "ignore").decode("ASCII")
    return texto


def normalizar_texto(texto):
    return remover_acentos(str(texto)).strip().lower()


def limpar_nome_arquivo(nome):
    nome = re.sub(r'[\\/:*?"<>|]', '', str(nome)).strip()
    return nome if nome else "arquivo.xlsx"


def limpar_nome_aba(nome):
    nome = os.path.splitext(str(nome))[0]
    nome = re.sub(r'[:\\/?*\[\]]', '', nome).strip()
    return (nome if nome else "Aba")[:31]


def formatar_duracao(segundos):
    segundos = int(max(0, segundos))
    h = segundos // 3600
    m = (segundos % 3600) // 60
    s = segundos % 60
    if h:
        return f"{h}h {m:02d}min {s:02d}s"
    if m:
        return f"{m}min {s:02d}s"
    return f"{s}s"


def calcular_tempos(inicio, concluidos, total):
    if concluidos <= 0 or total <= 0:
        return "calculando...", "calculando..."
    decorrido = time.time() - inicio
    restante = (decorrido / concluidos) * max(total - concluidos, 0)
    return formatar_duracao(decorrido), formatar_duracao(restante)


def normalizar_nome_para_mapeamento(nome):
    nome = os.path.splitext(str(nome))[0]
    nome = remover_acentos(nome).lower()
    nome = re.sub(r"[_\-.]+", " ", nome)
    nome = re.sub(r"\s+", " ", nome)
    return nome.strip()


def obter_nome_aba_final_personalizado(nome_arquivo_origem):
    """Mapeia nome do arquivo para a aba final padronizada."""
    nome_original_sem_ext = os.path.splitext(str(nome_arquivo_origem))[0]
    nome_sem_acento = remover_acentos(nome_original_sem_ext).lower().strip()
    nome_sem_acento = re.sub(r"\s+", " ", nome_sem_acento)
    nome_normalizado = normalizar_nome_para_mapeamento(nome_arquivo_origem)

    if "resumo gerencial cuiaba" in nome_normalizado:
        return "CUI"
    if "resumo gerencial diario ede" in nome_normalizado:
        return "EDE"
    if "rg sobradinho" in nome_normalizado:
        return "SOB"
    if re.search(r"(^|[^a-z0-9])gerencial_([0-2]?\d|3[01])([^0-9]|$)", nome_sem_acento):
        return "XAM"
    if re.search(r"\bgerencial\s+([0-2]\d|3[01])-[01]\d-\d{4}\b", nome_sem_acento):
        return "COB"
    if re.search(r"\bgerencial\s+([0-2]?\d|3[01])([^0-9-]|$)", nome_sem_acento):
        return "PVE"
    if "resumo gerencial" in nome_normalizado:
        return "NOB"

    nome_limpo = limpar_nome_aba(nome_arquivo_origem).upper()
    for aba in ORDEM_ABAS:
        if nome_limpo == aba or nome_limpo.startswith(aba + " ") or nome_limpo.startswith(aba + "_"):
            return aba
    return limpar_nome_aba(nome_arquivo_origem)


def identificar_aba_padrao(nome_aba):
    nome = limpar_nome_aba(nome_aba).upper().strip()
    for aba in ORDEM_ABAS:
        if nome == aba or nome.startswith(aba + " ") or nome.startswith(aba + "_") or nome.startswith(aba + "("):
            return aba
    return None


def aba_padrao_ja_existe(wb_final, nome_padrao):
    return any(identificar_aba_padrao(ws.title) == nome_padrao for ws in wb_final.worksheets)

# ============================================================
# DATAS: J3/K3 + NOME DA ABA
# ============================================================

def converter_para_date(valor):
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return None


def interpretar_data_em_texto(texto):
    if texto is None:
        return None
    texto = str(texto).strip()
    if not texto:
        return None
    t = remover_acentos(texto).lower().strip()
    t = re.sub(r"\s+", " ", t)

    # 06/06/2026, 06.06.2026, 06-06-2026
    m = re.search(r"\b([0-3]?\d)[\.\-/]([01]?\d)[\.\-/](\d{4})\b", t)
    if m:
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1))).date()
        except ValueError:
            pass

    # 6 de junho de 2026
    m = re.search(r"\b([0-3]?\d)\s+de\s+(janeiro|fevereiro|marco|março|abril|maio|junho|julho|agosto|setembro|outubro|novembro|dezembro)\s+de\s+(\d{4})\b", t)
    if m:
        try:
            dia = int(m.group(1))
            mes = MESES_PT.get(m.group(2), MESES_PT.get(remover_acentos(m.group(2))))
            ano = int(m.group(3))
            if mes:
                return datetime(ano, mes, dia).date()
        except ValueError:
            pass
    return None


def ler_data_j3_k3(aba):
    """Lê data em J3/K3. J3 é a principal, K3 ajuda se a data estiver separada."""
    valores = []
    for ref in ["J3", "K3"]:
        try:
            valor = aba[ref].value
        except Exception:
            valor = None
        data_valor = converter_para_date(valor)
        if data_valor:
            return data_valor
        if valor is not None:
            valores.append(str(valor))
    return interpretar_data_em_texto(" ".join(valores))


def interpretar_nome_aba_como_data(nome_aba, ano_referencia, mes_referencia):
    nome = str(nome_aba).strip()
    nome_norm = remover_acentos(nome).lower().strip()
    if nome_norm == normalizar_texto(NOME_ABA_GERENCIAL):
        return None

    apenas_digitos = re.sub(r"\D", "", nome_norm)
    try:
        if nome_norm.isdigit() and len(nome_norm) in [1, 2]:
            return datetime(ano_referencia, mes_referencia, int(nome_norm)).date()
        if apenas_digitos and len(apenas_digitos) == 4:
            dia = int(apenas_digitos[:2])
            mes = int(apenas_digitos[2:])
            return datetime(ano_referencia, mes, dia).date()
    except ValueError:
        return None
    return None


def obter_data_real_da_aba(aba, ano_referencia, mes_referencia):
    """
    Se a aba tiver data no nome e em J3/K3, as duas datas precisam bater.
    A aba Gerencial é validada pela data em J3/K3.
    """
    data_nome = interpretar_nome_aba_como_data(aba.title, ano_referencia, mes_referencia)
    data_j3 = ler_data_j3_k3(aba)

    if data_nome and data_j3:
        return data_nome if data_nome == data_j3 else None
    if data_nome:
        return data_nome
    if data_j3:
        return data_j3
    return None

# ============================================================
# CONSOLIDADO EXISTENTE
# ============================================================

def extrair_data_nome_arquivo(nome_arquivo):
    m = re.search(r"(\d{2})[\.\-/](\d{2})[\.\-/](\d{4})", str(nome_arquivo))
    if not m:
        return None
    try:
        return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1))).date()
    except ValueError:
        return None


def detectar_consolidado_existente(item):
    nome = remover_acentos(item["nome"]).lower()
    if "resumo gerencial cn" in nome or "consolidado" in nome:
        return True
    wb = None
    try:
        wb = load_workbook(item["caminho"], read_only=True, data_only=True, keep_links=False)
        return sum(1 for aba in wb.sheetnames if identificar_aba_padrao(aba) is not None) >= 2
    except Exception:
        return False
    finally:
        try:
            if wb:
                wb.close()
        except Exception:
            pass

# ============================================================
# ESCOLHA DA ABA PRINCIPAL — PONTO CRÍTICO
# ============================================================

def listar_registros_validos_do_arquivo(wb, ano_referencia, mes_referencia):
    registros = []
    for nome_aba in wb.sheetnames:
        aba = wb[nome_aba]
        data_real = obter_data_real_da_aba(aba, ano_referencia, mes_referencia)
        if data_real:
            registros.append({
                "aba": nome_aba,
                "data": data_real,
                "eh_gerencial": normalizar_texto(nome_aba) == normalizar_texto(NOME_ABA_GERENCIAL),
                "data_nome": interpretar_nome_aba_como_data(nome_aba, ano_referencia, mes_referencia),
            })
    return registros


def escolher_aba_principal_do_arquivo(wb, ano_referencia, mes_referencia):
    """
    Regra pedida:
    - Se existir aba Gerencial com data válida, usa SOMENTE a aba Gerencial desse arquivo.
      Não usa outras abas do mesmo arquivo, mesmo que tenham a mesma data.
    - Se não existir Gerencial com data válida, usa a maior data encontrada entre as outras abas.
    - Cada arquivo de origem entra em uma única data e gera no máximo uma aba final.
    """
    registros = listar_registros_validos_do_arquivo(wb, ano_referencia, mes_referencia)
    if not registros:
        return None

    gerenciais = [r for r in registros if r["eh_gerencial"]]
    if gerenciais:
        data_principal = max(r["data"] for r in gerenciais)
        candidatos = [r for r in gerenciais if r["data"] == data_principal]
        return candidatos[0]

    data_principal = max(r["data"] for r in registros)
    candidatos = [r for r in registros if r["data"] == data_principal]
    # Entre não Gerenciais, prioriza aba cujo nome indica a data.
    candidatos = sorted(candidatos, key=lambda r: 0 if r["data_nome"] == r["data"] else 1)
    return candidatos[0]


def gerar_chave_cache_datas(arquivos_salvos, ano_referencia, mes_referencia):
    assinatura = []
    for item in arquivos_salvos:
        caminho = item.get("caminho", "")
        nome = item.get("nome", "")
        tamanho = item.get("tamanho", 0)
        try:
            mtime = os.path.getmtime(caminho) if os.path.exists(caminho) else 0
        except Exception:
            mtime = 0
        assinatura.append((nome, tamanho, round(mtime, 3)))
    return (tuple(assinatura), ano_referencia, mes_referencia, "gerencial-tem-prioridade-exclusiva")


def construir_indice_datas(arquivos_salvos, ano_referencia, mes_referencia, quantidade_datas_filtro):
    chave = gerar_chave_cache_datas(arquivos_salvos, ano_referencia, mes_referencia)
    if (
        st.session_state.get("cache_datas_chave") == chave
        and "cache_datas_opcoes" in st.session_state
        and "cache_sem_data" in st.session_state
        and "cache_principal_arquivo" in st.session_state
    ):
        return (
            st.session_state.cache_datas_opcoes[:quantidade_datas_filtro],
            st.session_state.cache_principal_arquivo,
            st.session_state.cache_sem_data,
        )

    datas_encontradas = {}
    principal_arquivo = {}
    sem_data = []
    inicio = time.time()
    status = st.empty()
    barra = st.progress(0)
    total = max(len(arquivos_salvos), 1)

    for idx, item in enumerate(arquivos_salvos, start=1):
        wb = None
        caminho = item["caminho"]
        try:
            status.info(f"Lendo data principal: {idx}/{len(arquivos_salvos)} — {item['nome']}")
            wb = load_workbook(caminho, read_only=True, data_only=True, keep_links=False)

            if detectar_consolidado_existente(item):
                data_nome = extrair_data_nome_arquivo(item["nome"])
                if data_nome:
                    principal_arquivo[caminho] = {"data": data_nome, "aba": None, "tipo": "consolidado"}
                    datas_encontradas.setdefault(data_nome, {"data": datetime.combine(data_nome, datetime.min.time()), "arquivos": [], "abas": []})
                    datas_encontradas[data_nome]["arquivos"].append(item["nome"])
                    datas_encontradas[data_nome]["abas"].extend([a for a in wb.sheetnames if identificar_aba_padrao(a)])
                    continue

            registro = escolher_aba_principal_do_arquivo(wb, ano_referencia, mes_referencia)
            if not registro:
                sem_data.append({"arquivo": item["nome"], "motivo": "Nenhuma data identificável por nome da aba ou J3/K3"})
                continue

            data_principal = registro["data"]
            principal_arquivo[caminho] = {"data": data_principal, "aba": registro["aba"], "tipo": "origem", "eh_gerencial": registro["eh_gerencial"]}
            datas_encontradas.setdefault(data_principal, {"data": datetime.combine(data_principal, datetime.min.time()), "arquivos": [], "abas": []})
            datas_encontradas[data_principal]["arquivos"].append(item["nome"])
            datas_encontradas[data_principal]["abas"].append(registro["aba"])

        except Exception as erro:
            sem_data.append({"arquivo": item["nome"], "motivo": f"Erro ao ler arquivo: {erro}"})
        finally:
            try:
                if wb:
                    wb.close()
            except Exception:
                pass
            gc.collect()
            barra.progress(idx / total)

    decorrido = formatar_duracao(time.time() - inicio)
    status.success(f"Leitura de datas concluída em {decorrido}.")
    lista_datas = sorted(datas_encontradas.values(), key=lambda item: item["data"], reverse=True)

    st.session_state.cache_datas_chave = chave
    st.session_state.cache_datas_opcoes = lista_datas
    st.session_state.cache_principal_arquivo = principal_arquivo
    st.session_state.cache_sem_data = sem_data
    return lista_datas[:quantidade_datas_filtro], principal_arquivo, sem_data

# ============================================================
# UPLOAD / SESSÃO
# ============================================================

def criar_pasta_sessao():
    if "pasta_sessao" not in st.session_state:
        st.session_state.pasta_sessao = tempfile.mkdtemp()


def limpar_cache_datas():
    for chave in ["cache_datas_chave", "cache_datas_opcoes", "cache_sem_data", "cache_principal_arquivo"]:
        st.session_state.pop(chave, None)


def salvar_uploads_em_pasta_sessao(arquivos_upload):
    criar_pasta_sessao()
    arquivos_salvos = []
    for uploaded_file in arquivos_upload:
        nome_limpo = limpar_nome_arquivo(uploaded_file.name)
        caminho_base = os.path.join(st.session_state.pasta_sessao, nome_limpo)
        caminho_final = caminho_base
        contador = 1
        while os.path.exists(caminho_final):
            nome_sem_extensao, extensao = os.path.splitext(nome_limpo)
            caminho_final = os.path.join(st.session_state.pasta_sessao, f"{nome_sem_extensao}_{contador}{extensao}")
            contador += 1
        with open(caminho_final, "wb") as f:
            f.write(uploaded_file.getvalue())
        arquivos_salvos.append({"nome": uploaded_file.name, "caminho": caminho_final, "tamanho": uploaded_file.size})
    limpar_cache_datas()
    return arquivos_salvos


def limpar_arquivos_da_sessao():
    if "pasta_sessao" in st.session_state:
        try:
            if os.path.exists(st.session_state.pasta_sessao):
                shutil.rmtree(st.session_state.pasta_sessao)
        except Exception:
            pass
    st.session_state.arquivos_salvos = []
    st.session_state.uploader_key += 1
    limpar_cache_datas()
    if "pasta_sessao" in st.session_state:
        del st.session_state.pasta_sessao

# ============================================================
# FORMATAÇÃO / CÓPIA
# ============================================================

def aplicar_configuracoes_finais_aba(aba_destino):
    aba_destino.sheet_view.zoomScale = 80
    aba_destino.sheet_view.zoomScaleNormal = 80
    aba_destino.freeze_panes = "A5"
    aba_destino.sheet_view.showGridLines = False


def limpar_conteudos_a_partir_da_coluna_m(aba_destino):
    if aba_destino.max_column < COLUNA_INICIO_LIMPEZA:
        return
    for intervalo in list(aba_destino.merged_cells.ranges):
        if intervalo.min_col >= COLUNA_INICIO_LIMPEZA:
            cel = aba_destino.cell(row=intervalo.min_row, column=intervalo.min_col)
            if not isinstance(cel, MergedCell):
                cel.value = None
                cel.comment = None
                cel._hyperlink = None
    for linha in aba_destino.iter_rows(min_row=1, max_row=aba_destino.max_row, min_col=COLUNA_INICIO_LIMPEZA, max_col=aba_destino.max_column):
        for cel in linha:
            if isinstance(cel, MergedCell):
                continue
            cel.value = None
            cel.comment = None
            cel._hyperlink = None


def garantir_requisitos_todas_abas(wb_final):
    for aba in wb_final.worksheets:
        limpar_conteudos_a_partir_da_coluna_m(aba)
        aplicar_configuracoes_finais_aba(aba)


def copiar_aba_como_valores_mais_leve(aba_origem, aba_destino, copiar_estilos=True):
    for linha in aba_origem.iter_rows(min_row=1, max_row=aba_origem.max_row, min_col=1, max_col=aba_origem.max_column):
        for celula_origem in linha:
            if isinstance(celula_origem, MergedCell):
                continue
            celula_destino = aba_destino[celula_origem.coordinate]
            celula_destino.value = celula_origem.value
            if copiar_estilos and celula_origem.has_style:
                celula_destino.font = copy(celula_origem.font)
                celula_destino.fill = copy(celula_origem.fill)
                celula_destino.border = copy(celula_origem.border)
                celula_destino.alignment = copy(celula_origem.alignment)
                celula_destino.number_format = celula_origem.number_format
                celula_destino.protection = copy(celula_origem.protection)
            if celula_origem.comment:
                celula_destino.comment = copy(celula_origem.comment)
            if celula_origem.hyperlink:
                celula_destino._hyperlink = copy(celula_origem.hyperlink)

    for letra, dim in aba_origem.column_dimensions.items():
        aba_destino.column_dimensions[letra].width = dim.width
        aba_destino.column_dimensions[letra].hidden = dim.hidden
        aba_destino.column_dimensions[letra].outlineLevel = dim.outlineLevel
        aba_destino.column_dimensions[letra].collapsed = dim.collapsed
    for numero, dim in aba_origem.row_dimensions.items():
        aba_destino.row_dimensions[numero].height = dim.height
        aba_destino.row_dimensions[numero].hidden = dim.hidden
        aba_destino.row_dimensions[numero].outlineLevel = dim.outlineLevel
        aba_destino.row_dimensions[numero].collapsed = dim.collapsed
    for intervalo in aba_origem.merged_cells.ranges:
        try:
            aba_destino.merge_cells(str(intervalo))
        except Exception:
            pass
    if aba_origem.auto_filter and aba_origem.auto_filter.ref:
        aba_destino.auto_filter.ref = aba_origem.auto_filter.ref


def padronizar_larguras_colunas_pela_aba_ede(wb_final, nome_aba_referencia="EDE"):
    if nome_aba_referencia not in wb_final.sheetnames:
        return False, "A aba de referência 'EDE' não foi encontrada. As larguras não foram padronizadas."
    aba_referencia = wb_final[nome_aba_referencia]
    max_colunas = max(ws.max_column for ws in wb_final.worksheets)
    for idx in range(1, max_colunas + 1):
        letra = aba_referencia.cell(row=1, column=idx).column_letter
        largura = aba_referencia.column_dimensions[letra].width
        if largura is None:
            largura = aba_referencia.sheet_format.defaultColWidth or 8.43
        for aba in wb_final.worksheets:
            aba.column_dimensions[letra].width = largura
    return True, "Larguras das colunas padronizadas com base na aba 'EDE'."


def ordenar_abas(wb_final):
    def prioridade(ws):
        nome_padrao = identificar_aba_padrao(ws.title)
        if nome_padrao in ORDEM_ABAS:
            return ORDEM_ABAS.index(nome_padrao)
        return 99
    wb_final._sheets = sorted(wb_final.worksheets, key=prioridade)

# ============================================================
# PROCESSAMENTO MULTIDATA
# ============================================================

def copiar_abas_de_consolidado_existente(wb_origem, wb_final, data_processar, ano_ref, mes_ref, logs, nome_arquivo):
    copiadas = []
    data_nome_arquivo = extrair_data_nome_arquivo(nome_arquivo)
    for nome_aba in wb_origem.sheetnames:
        nome_padrao = identificar_aba_padrao(nome_aba)
        if nome_padrao not in ORDEM_ABAS:
            continue
        if aba_padrao_ja_existe(wb_final, nome_padrao):
            logs.append(f"Consolidado existente: aba '{nome_padrao}' ignorada porque já existe.")
            continue
        aba_origem = wb_origem[nome_aba]
        data_aba = obter_data_real_da_aba(aba_origem, ano_ref, mes_ref)
        if data_aba is not None and data_aba != data_processar:
            continue
        if data_aba is None and data_nome_arquivo != data_processar:
            continue
        aba_destino = wb_final.create_sheet(title=nome_padrao)
        copiar_aba_como_valores_mais_leve(aba_origem, aba_destino, copiar_estilos=True)
        limpar_conteudos_a_partir_da_coluna_m(aba_destino)
        aplicar_configuracoes_finais_aba(aba_destino)
        copiadas.append(nome_padrao)
        logs.append(f"Consolidado existente '{nome_arquivo}': aba '{nome_aba}' copiada como '{nome_padrao}'.")
    return copiadas


def processar_data(data_processar, arquivos_salvos, principal_arquivo, data_referencia, copiar_estilos=True, callback_progresso=None):
    logs, arquivos_processados, arquivos_com_erro, abas_criadas = [], [], [], []
    ano_ref = data_referencia.year
    mes_ref = data_referencia.month
    nome_arquivo_final = f"Resumo Gerencial CN - {data_processar.strftime('%d.%m.%Y')}.xlsx"
    wb_final = Workbook()
    wb_final.remove(wb_final.active)

    logs.append(f"Data do consolidado: {data_processar.strftime('%d/%m/%Y')}")
    logs.append("Regra: se o arquivo tiver aba Gerencial com data válida, copia somente essa Gerencial. Não copia outras abas do mesmo arquivo.")
    logs.append("Regra: cada arquivo de origem entra somente na data principal dele. Não mistura datas no consolidado.")

    # 1) Consolidados existentes primeiro.
    for item in arquivos_salvos:
        if not detectar_consolidado_existente(item):
            continue
        wb = None
        try:
            info = principal_arquivo.get(item["caminho"])
            data_principal = info["data"] if info else extrair_data_nome_arquivo(item["nome"])
            if data_principal != data_processar:
                if callback_progresso:
                    callback_progresso(item["nome"], data_processar)
                continue
            wb = load_workbook(item["caminho"], data_only=True, keep_links=False)
            copiadas = copiar_abas_de_consolidado_existente(wb, wb_final, data_processar, ano_ref, mes_ref, logs, item["nome"])
            if copiadas:
                abas_criadas.extend(copiadas)
                arquivos_processados.append({"arquivo": item["nome"], "aba_origem": "Consolidado existente", "aba_final": ", ".join(copiadas), "criterio": "Reaproveitado de consolidado parcial da mesma data"})
        except Exception as erro:
            arquivos_com_erro.append({"arquivo": item["nome"], "erro": str(erro)})
        finally:
            try:
                if wb:
                    wb.close()
            except Exception:
                pass
            gc.collect()
            if callback_progresso:
                callback_progresso(item["nome"], data_processar)

    # 2) Planilhas de origem/faltantes.
    for item in arquivos_salvos:
        if detectar_consolidado_existente(item):
            continue
        wb = None
        try:
            info = principal_arquivo.get(item["caminho"])
            if not info or info.get("data") != data_processar:
                if callback_progresso:
                    callback_progresso(item["nome"], data_processar)
                continue

            nome_padrao = obter_nome_aba_final_personalizado(item["nome"])
            nome_padrao = identificar_aba_padrao(nome_padrao) or nome_padrao
            if nome_padrao not in ORDEM_ABAS:
                logs.append(f"IGNORADO: {item['nome']} não mapeia para uma das abas padrão.")
                if callback_progresso:
                    callback_progresso(item["nome"], data_processar)
                continue
            if aba_padrao_ja_existe(wb_final, nome_padrao):
                logs.append(f"IGNORADO: {item['nome']} porque a aba '{nome_padrao}' já existe no consolidado da data.")
                if callback_progresso:
                    callback_progresso(item["nome"], data_processar)
                continue

            wb = load_workbook(item["caminho"], data_only=True, keep_links=False)
            nome_aba_origem = info.get("aba")
            if not nome_aba_origem or nome_aba_origem not in wb.sheetnames:
                logs.append(f"IGNORADO: {item['nome']} não teve aba principal identificada.")
                if callback_progresso:
                    callback_progresso(item["nome"], data_processar)
                continue

            aba_origem = wb[nome_aba_origem]
            data_confirmada = obter_data_real_da_aba(aba_origem, ano_ref, mes_ref)
            if data_confirmada != data_processar:
                logs.append(f"IGNORADO: {item['nome']} aba '{nome_aba_origem}' tem data {data_confirmada}, não {data_processar}.")
                if callback_progresso:
                    callback_progresso(item["nome"], data_processar)
                continue

            aba_destino = wb_final.create_sheet(title=nome_padrao)
            copiar_aba_como_valores_mais_leve(aba_origem, aba_destino, copiar_estilos=copiar_estilos)
            limpar_conteudos_a_partir_da_coluna_m(aba_destino)
            aplicar_configuracoes_finais_aba(aba_destino)
            criterio = "Gerencial exclusiva do arquivo" if info.get("eh_gerencial") else "Aba não Gerencial principal do arquivo"
            arquivos_processados.append({"arquivo": item["nome"], "aba_origem": nome_aba_origem, "aba_final": nome_padrao, "criterio": f"{criterio}; data = {data_processar.strftime('%d/%m/%Y')}"})
            abas_criadas.append(nome_padrao)
            logs.append(f"SUCESSO: {item['nome']} | '{nome_aba_origem}' -> '{nome_padrao}'.")
        except Exception as erro:
            arquivos_com_erro.append({"arquivo": item.get("nome", "arquivo"), "erro": str(erro)})
            logs.append(f"ERRO: {item.get('nome', 'arquivo')} - {erro}")
        finally:
            try:
                if wb:
                    wb.close()
            except Exception:
                pass
            gc.collect()
            if callback_progresso:
                callback_progresso(item.get("nome", "arquivo"), data_processar)

    if not wb_final.sheetnames:
        raise Exception(f"Nenhuma aba foi criada para {data_processar.strftime('%d/%m/%Y')}.")

    garantir_requisitos_todas_abas(wb_final)
    _, msg_larguras = padronizar_larguras_colunas_pela_aba_ede(wb_final, "EDE")
    logs.append(msg_larguras)
    garantir_requisitos_todas_abas(wb_final)
    ordenar_abas(wb_final)

    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
    wb_final.save(temp)
    wb_final.close()
    with open(temp, "rb") as f:
        conteudo = f.read()
    try:
        os.remove(temp)
    except Exception:
        pass
    output = BytesIO(conteudo)
    output.seek(0)
    return {"arquivo_excel": output, "nome_arquivo_final": nome_arquivo_final, "logs": logs, "arquivos_processados": arquivos_processados, "arquivos_com_erro": arquivos_com_erro, "abas_criadas": abas_criadas, "data_processada": data_processar}


def processar_multiplas_datas(datas_selecionadas, arquivos_salvos, principal_arquivo, data_referencia, copiar_estilos=True):
    resultados = []
    datas_ordenadas = sorted([item["data"] for item in datas_selecionadas], reverse=True)
    total_etapas = max(len(datas_ordenadas) * len(arquivos_salvos), 1)
    concluidos = 0
    inicio = time.time()
    barra = st.progress(0)
    status = st.empty()
    tempo_box = st.empty()

    def progresso(nome_arquivo, data_atual):
        nonlocal concluidos
        concluidos += 1
        barra.progress(min(concluidos / total_etapas, 1))
        decorrido, restante = calcular_tempos(inicio, concluidos, total_etapas)
        tempo_box.info(f"⏱️ Tempo decorrido: {decorrido} | Tempo restante estimado: {restante} | Progresso: {concluidos}/{total_etapas} etapa(s)")
        status.info(f"Processando {data_atual.strftime('%d/%m/%Y')} | Arquivo: {nome_arquivo}")

    for idx, data in enumerate(datas_ordenadas, start=1):
        status.info(f"Gerando consolidado {idx}/{len(datas_ordenadas)}: {data.strftime('%d/%m/%Y')}")
        resultado = processar_data(data, arquivos_salvos, principal_arquivo, data_referencia, copiar_estilos=copiar_estilos, callback_progresso=progresso)
        resultados.append(resultado)

    tempo_total = formatar_duracao(time.time() - inicio)
    tempo_box.success(f"✅ Processamento concluído em {tempo_total}.")
    status.success("Todos os consolidados selecionados foram gerados.")
    return resultados


def gerar_zip_resultados(resultados):
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as z:
        for resultado in resultados:
            z.writestr(resultado["nome_arquivo_final"], resultado["arquivo_excel"].getvalue())
    zip_buffer.seek(0)
    return zip_buffer

# ============================================================
# INTERFACE
# ============================================================

if "arquivos_salvos" not in st.session_state:
    st.session_state.arquivos_salvos = []
if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0

st.title("📊 Sistema de Consolidação Gerencial CN")
st.markdown(
    """
    Este sistema consolida arquivos Excel por data.

    **Regras principais:**
    - cada consolidado contém somente abas da data selecionada;
    - se um arquivo tiver aba `Gerencial` com data válida, o sistema copia **somente a Gerencial desse arquivo**;
    - se o arquivo não tiver Gerencial válida, o sistema usa a aba não Gerencial principal da data;
    - arquivos com histórico interno entram somente na data principal do próprio arquivo;
    - se você enviar 21 planilhas de 3 datas diferentes, o sistema gera até 3 consolidados com até 7 abas cada;
    - também é possível enviar um consolidado parcial junto com planilhas faltantes.
    """
)
st.divider()

st.sidebar.header("⚙️ Configurações")
usar_data_ontem = st.sidebar.checkbox("Usar ontem como referência de mês/ano", value=True)
if usar_data_ontem:
    data_referencia = datetime.today() - timedelta(days=1)
    st.sidebar.info(f"Referência: {data_referencia.strftime('%m/%Y')}")
else:
    data_escolhida = st.sidebar.date_input("Escolha uma data de referência para mês/ano", value=(datetime.today() - timedelta(days=1)).date())
    data_referencia = datetime.combine(data_escolhida, datetime.min.time())

qtde_esperada = st.sidebar.number_input("Quantidade esperada de arquivos", min_value=1, max_value=100, value=7, step=1)
copiar_estilos = st.sidebar.checkbox("Copiar formatação das células", value=True)
quantidade_datas_filtro = st.sidebar.number_input("Quantidade máxima de datas no filtro", min_value=1, max_value=31, value=5, step=1)
max_datas_processar = st.sidebar.number_input("Máximo de datas para processar de uma vez", min_value=1, max_value=31, value=5, step=1)

st.subheader("1. Envie os arquivos Excel")
arquivos_upload = st.file_uploader("Selecione todos os arquivos Excel de uma vez. Você pode incluir o consolidado parcial e as planilhas faltantes.", type=["xlsx", "xlsm"], accept_multiple_files=True, key=f"uploader_{st.session_state.uploader_key}")

col_carregar, col_limpar = st.columns(2)
with col_carregar:
    if st.button("📥 Carregar arquivos enviados", disabled=not arquivos_upload):
        try:
            limpar_arquivos_da_sessao()
            st.session_state.arquivos_salvos = salvar_uploads_em_pasta_sessao(arquivos_upload)
            st.success(f"{len(st.session_state.arquivos_salvos)} arquivo(s) carregado(s) com sucesso.")
            st.rerun()
        except Exception as erro:
            st.error(f"Erro ao carregar arquivos: {erro}")
with col_limpar:
    if st.button("🧹 Limpar arquivos carregados"):
        limpar_arquivos_da_sessao()
        st.success("Arquivos carregados foram limpos.")
        st.rerun()

arquivos_salvos = st.session_state.arquivos_salvos
if arquivos_salvos:
    st.success(f"{len(arquivos_salvos)} arquivo(s) carregado(s).")
    if len(arquivos_salvos) != qtde_esperada:
        st.warning(f"Foram carregados {len(arquivos_salvos)} arquivo(s), mas a quantidade esperada é {qtde_esperada}.")
    tamanho_total_mb = sum(item["tamanho"] for item in arquivos_salvos) / (1024 * 1024)
    st.info(f"Tamanho total carregado: {tamanho_total_mb:.2f} MB")
    with st.expander("Ver arquivos carregados"):
        for i, item in enumerate(arquivos_salvos, start=1):
            nome_padronizado = obter_nome_aba_final_personalizado(item["nome"])
            tipo = "Consolidado existente" if detectar_consolidado_existente(item) else "Planilha de origem/faltante"
            st.write(f"{i}. {item['nome']} — tipo: {tipo} — aba final prevista: {nome_padronizado}")
else:
    st.info("Envie os arquivos Excel e clique em 'Carregar arquivos enviados'.")

st.divider()
st.subheader("2. Escolha uma ou mais datas")
datas_selecionadas = []
principal_arquivo = {}

if arquivos_salvos:
    datas_para_filtro, principal_arquivo, sem_data = construir_indice_datas(arquivos_salvos, data_referencia.year, data_referencia.month, quantidade_datas_filtro)
    if datas_para_filtro:
        opcoes_datas = []
        for item in datas_para_filtro:
            data_item = item["data"].date()
            qtd_arquivos = len(set(item["arquivos"]))
            label = f"{data_item.strftime('%d/%m/%Y')} — data principal em {qtd_arquivos} arquivo(s)"
            opcoes_datas.append({"label": label, "data": data_item})
        datas_selecionadas = st.multiselect("Selecione uma ou mais datas", options=opcoes_datas, default=opcoes_datas[:1], format_func=lambda item: item["label"])
        if datas_selecionadas:
            st.success(f"{len(datas_selecionadas)} data(s) selecionada(s).")
    else:
        st.warning("Nenhuma data foi encontrada nas abas dos arquivos carregados.")
    if sem_data:
        with st.expander("Arquivos sem data identificável"):
            for item in sem_data[:200]:
                st.write(f"- **Arquivo:** {item['arquivo']} | **Motivo:** {item['motivo']}")
else:
    st.info("Carregue os arquivos para o sistema listar as datas disponíveis.")

st.divider()
st.subheader("3. Processar")
botao_processar = st.button("🚀 Gerar consolidado(s) selecionado(s)", type="primary", disabled=len(st.session_state.arquivos_salvos) == 0 or len(datas_selecionadas) == 0)

if botao_processar:
    try:
        if len(datas_selecionadas) > max_datas_processar:
            st.warning(f"Selecione no máximo {max_datas_processar} data(s) por execução.")
            st.stop()
        with st.spinner("Processando arquivos..."):
            resultados = processar_multiplas_datas(datas_selecionadas, arquivos_salvos, principal_arquivo, data_referencia, copiar_estilos=copiar_estilos)

        st.success("Arquivo(s) final(is) gerado(s) com sucesso!")
        total_processados = sum(len(r["arquivos_processados"]) for r in resultados)
        total_erros = sum(len(r["arquivos_com_erro"]) for r in resultados)
        col1, col2, col3 = st.columns(3)
        col1.metric("Consolidados gerados", len(resultados))
        col2.metric("Abas/arquivos processados", total_processados)
        col3.metric("Com erro", total_erros)

        st.subheader("4. Downloads")
        if len(resultados) > 1:
            zip_resultados = gerar_zip_resultados(resultados)
            st.download_button("⬇️ Baixar todos os consolidados em ZIP", zip_resultados, "Consolidados_Gerenciais_CN.zip", mime="application/zip")
        for resultado in resultados:
            st.download_button(f"⬇️ Baixar {resultado['nome_arquivo_final']}", resultado["arquivo_excel"], resultado["nome_arquivo_final"], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        st.subheader("5. Resumo")
        for resultado in resultados:
            data_proc = resultado["data_processada"].strftime("%d/%m/%Y")
            with st.expander(f"Resumo do consolidado da data {data_proc}"):
                if resultado["arquivos_processados"]:
                    st.markdown("### ✅ Arquivos processados")
                    for item in resultado["arquivos_processados"]:
                        st.write(f"- **Arquivo:** {item['arquivo']} | **Aba origem:** {item['aba_origem']} | **Aba final:** {item['aba_final']} | **Critério:** {item['criterio']}")
                if resultado["abas_criadas"]:
                    st.markdown("### Abas criadas")
                    st.write(", ".join(resultado["abas_criadas"]))
                if resultado["arquivos_com_erro"]:
                    st.markdown("### ❌ Arquivos com erro")
                    for item in resultado["arquivos_com_erro"]:
                        st.error(f"{item['arquivo']}: {item['erro']}")
                st.markdown("### Logs")
                st.code("\n".join(resultado["logs"]), language="text")
    except Exception as erro:
        st.error(f"Erro geral ao processar os arquivos: {erro}")
        st.code(traceback.format_exc(), language="text")

st.divider()
st.caption("Cada consolidado contém apenas abas da data selecionada. Se uma planilha tiver aba Gerencial válida, somente a Gerencial será copiada daquele arquivo. Ordem final: COB, CUI, EDE, NOB, PVE, SOB, XAM.")
